from flask import Flask, render_template, request, send_from_directory, send_file, redirect, url_for, session
from werkzeug.security import check_password_hash, generate_password_hash
import pandas as pd
import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

app = Flask(__name__)

current_date = datetime.now().strftime("%Y-%m-%d")
OUTPUT_FOLDER = "outputs"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
#################### 2nd floor syn ###################################################
def process_2nd_floor(file_path):

    df = pd.read_excel(file_path)

    required_columns = ["Employee Code", "Employee Name", "Date", "Door Name.", "Time"]
    df_filtered = df[required_columns]

    df_filtered["Date"] = pd.to_datetime(df_filtered["Date"]).dt.strftime("%Y-%m-%d")
    df_filtered["Time"] = pd.to_datetime(df_filtered["Time"], format="%H:%M:%S").dt.time

    df_in = df_filtered[df_filtered["Door Name."].str.contains("Access Control Reader 1", na=False)].copy()
    df_out = df_filtered[df_filtered["Door Name."].str.contains("Access Control Reader 2", na=False)].copy()

    df_in_grouped = df_in.groupby(["Employee Code", "Employee Name", "Date"])["Time"].apply(list).reset_index()
    df_out_grouped = df_out.groupby(["Employee Code", "Employee Name", "Date"])["Time"].apply(list).reset_index()

    df_merged = pd.merge(df_in_grouped, df_out_grouped, on=["Employee Code", "Employee Name", "Date"], how="outer", suffixes=("_IN", "_OUT"))

    df_merged["Time_IN"] = df_merged["Time_IN"].apply(lambda x: sorted(x) if isinstance(x, list) else [])
    df_merged["Time_OUT"] = df_merged["Time_OUT"].apply(lambda x: sorted(x) if isinstance(x, list) else [])

    def pair_in_out(in_times, out_times):
        paired = []
        total_duration = timedelta()
        in_index, out_index = 0, 0

        while in_index < len(in_times) and out_index < len(out_times):
            while in_index < len(in_times) - 1 and in_times[in_index + 1] < out_times[out_index]:
                in_index += 1
            while out_index < len(out_times) - 1 and (in_index + 1 >= len(in_times) or out_times[out_index + 1] < in_times[in_index + 1]):
                out_index += 1

            if in_index < len(in_times) and out_index < len(out_times) and out_times[out_index] > in_times[in_index]:
                duration = (datetime.combine(datetime.today(), out_times[out_index]) - datetime.combine(datetime.today(), in_times[in_index])).seconds
                total_duration += timedelta(seconds=duration)
                paired.append((in_times[in_index], out_times[out_index]))
                in_index += 1
                out_index += 1
            else:
                out_index += 1

        return paired, total_duration

    expanded_data = []
    for _, row in df_merged.iterrows():
        paired_times, total_duration = pair_in_out(sorted(row["Time_IN"]), sorted(row["Time_OUT"]))

        row_data = {
            "Employee Code": row["Employee Code"],
            "Employee Name": row["Employee Name"],
            "Date": row["Date"],
            "Total Time Spent": str(total_duration).split()[-1] if total_duration else "00:00:00"
        }

        for i, (in_time, out_time) in enumerate(paired_times):
            row_data[f"IN {i+1}"] = in_time
            row_data[f"OUT {i+1}"] = out_time

        expanded_data.append(row_data)

    df_final = pd.DataFrame(expanded_data)
    output_path = os.path.join(OUTPUT_FOLDER, "2nd_floor.xlsx")
    df_final.to_excel(output_path, index=False)
    return output_path
#################### 3rd floor syn ###################################################
def process_3rd_floor(file_path):
    def calculate_total_time(punch):
        if pd.isna(punch) or punch.strip() == "":
            return "00:00:00"

        punch_times = re.findall(r'(\d{2}:\d{2})\((in|out)\)', punch)
        events = [(datetime.strptime(time, "%H:%M"), status) for time, status in punch_times]

        total_time = 0
        i = 0

        # Sort events to ensure they are in correct chronological order
        events.sort(key=lambda x: x[0])

        while i < len(events):
            if events[i][1] == "in":
                in_time = events[i][0]
                i += 1

                # Find the last out before next in
                last_out = None
                while i < len(events) and events[i][1] == "out":
                    last_out = events[i][0]
                    i += 1

                if last_out:
                    total_time += (last_out - in_time).seconds
            else:
                i += 1  # Skip unmatched 'out'

        return str(timedelta(seconds=total_time))

    df = pd.read_excel(file_path)
    df.columns = df.columns.str.strip()
    df_selected = df[["Employee Code", "Employee Name", "Date", "Punch Records"]].copy()
    df_selected["Date"] = pd.to_datetime(df_selected["Date"]).dt.date

    # Apply updated logic
    total_times = []

    for punch in df_selected["Punch Records"]:
        total_time = calculate_total_time(punch)
        total_times.append(total_time)

    df_selected["Total Time Spent"] = total_times

    output_path = os.path.join(OUTPUT_FOLDER, "3rd_floor.xlsx")
    df_selected.to_excel(output_path, index=False)
    return output_path


#################### 2nd & 3rd floor syn ###################################################
def merge_files(file_2nd, file_3rd):

    df_2nd = pd.read_excel(file_2nd)
    df_3rd = pd.read_excel(file_3rd)

    df_2nd_filtered = df_2nd[['Employee Code', 'Total Time Spent']].rename(columns={'Total Time Spent': 'Total Time Spent 2nd Floor'})
    df_3rd_filtered = df_3rd[['Employee Code', 'Total Time Spent']].rename(columns={'Total Time Spent': 'Total Time Spent 3rd Floor'})

    merged_df = pd.merge(df_2nd_filtered, df_3rd_filtered, on='Employee Code', how='inner')

    merged_df['Total Time Spent 2nd Floor'] = pd.to_timedelta(merged_df['Total Time Spent 2nd Floor'])
    merged_df['Total Time Spent 3rd Floor'] = pd.to_timedelta(merged_df['Total Time Spent 3rd Floor'])

    merged_df['Total Time Spent 2nd & 3rd Floor'] = merged_df['Total Time Spent 2nd Floor'] + merged_df['Total Time Spent 3rd Floor']

    merged_df['Total Time Spent 2nd Floor'] = merged_df['Total Time Spent 2nd Floor'].apply(lambda x: str(x).split()[-1])
    merged_df['Total Time Spent 3rd Floor'] = merged_df['Total Time Spent 3rd Floor'].apply(lambda x: str(x).split()[-1])
    merged_df['Total Time Spent 2nd & 3rd Floor'] = merged_df['Total Time Spent 2nd & 3rd Floor'].apply(lambda x: str(x).split()[-1])

    output_path = os.path.join(OUTPUT_FOLDER, f"2nd_floor&3rd_floor_Combine_{current_date}.xlsx")
    merged_df.to_excel(output_path, index=False)
    return output_path

@app.route("/2-3", methods=["GET", "POST"])
def upload_files():
    if request.method == "POST":
        file_2nd = request.files["file_2nd"]
        file_3rd = request.files["file_3rd"]

        path_2nd = process_2nd_floor(file_2nd)
        path_3rd = process_3rd_floor(file_3rd)
        merged_path = merge_files(path_2nd, path_3rd)

        return render_template("upload.html", processed=True)

    return render_template("upload.html", processed=False)

@app.route("/download/<filename>")
def download_file(filename):
    return send_from_directory(OUTPUT_FOLDER, filename, as_attachment=True)


#####################################################################################

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Function to convert time string to total seconds
def time_to_seconds(time_str):
    try:
        h, m, s = map(int, time_str.split(":"))
        return h * 3600 + m * 60 + s
    except:
        return 0  # Return 0 if the value is not in time format

# Function to convert total seconds back to HH:MM:SS format
def seconds_to_time(seconds):
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02}:{m:02}:{s:02}"

@app.route("/merged", methods=["GET", "POST"])
def merged():
    if request.method == "POST":
        # Get uploaded files
        file_2nd = request.files.get("file_2nd")
        file_3rd = request.files.get("file_3rd")

        if not file_2nd or not file_3rd:
            return "Please upload both Excel files."

        # Save uploaded files
        file_2nd_path = os.path.join(UPLOAD_FOLDER, file_2nd.filename)
        file_3rd_path = os.path.join(UPLOAD_FOLDER, file_3rd.filename)
        file_2nd.save(file_2nd_path)
        file_3rd.save(file_3rd_path)

        # Load the Excel files
        df_2nd = pd.read_excel(file_2nd_path)
        df_3rd = pd.read_excel(file_3rd_path)

        # Process the data
        df_2nd = df_2nd.rename(columns={"Total Time Spent": "Total Time Spent 2nd Floor"})
        df_3rd = df_3rd.rename(columns={"Total Time Spent": "Total Time Spent 3rd Floor"})

        df_2nd = df_2nd[["Date", "Employee Code", "Employee Name", "Total Time Spent 2nd Floor"]]
        df_3rd = df_3rd[["Date", "Employee Code", "Employee Name", "Total Time Spent 3rd Floor"]]

        df_2nd["Date"] = pd.to_datetime(df_2nd["Date"], errors="coerce")
        df_3rd["Date"] = pd.to_datetime(df_3rd["Date"], errors="coerce")

        df_merged = pd.merge(
            df_2nd, df_3rd, on=["Date", "Employee Code"], how="outer", suffixes=("_2nd", "_3rd")
        )
        df_merged["Employee Name"] = df_merged["Employee Name_2nd"].combine_first(df_merged["Employee Name_3rd"])

        df_merged = df_merged[
            ["Date", "Employee Code", "Employee Name", "Total Time Spent 2nd Floor", "Total Time Spent 3rd Floor"]
        ]
        df_merged.fillna({"Total Time Spent 2nd Floor": "00:00:00", "Total Time Spent 3rd Floor": "00:00:00"}, inplace=True)
        df_merged["Date"] = df_merged["Date"].dt.strftime("%Y-%m-%d")
        df_merged["Seconds_2nd"] = df_merged["Total Time Spent 2nd Floor"].apply(time_to_seconds)
        df_merged["Seconds_3rd"] = df_merged["Total Time Spent 3rd Floor"].apply(time_to_seconds)

        df_merged["Total Time Spent (Both Floors)"] = df_merged["Seconds_2nd"] + df_merged["Seconds_3rd"]
        df_merged["Total Time Spent (Both Floors)"] = df_merged["Total Time Spent (Both Floors)"].apply(seconds_to_time)
        df_merged.drop(columns=["Seconds_2nd", "Seconds_3rd"], inplace=True)

        output_file = os.path.join(OUTPUT_FOLDER, f"Merged_Floor_{current_date}.xlsx")
        df_merged.to_excel(output_file, index=False)

        # Apply coloring and Excel-style remark formula
        apply_remark_coloring(output_file)

        return send_file(output_file, as_attachment=True)

    return render_template("merged.html")

# ---------- Helper Function to Apply Remark + Conditional Formatting ----------
def apply_remark_coloring(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    time_col_idx = headers.index("Total Time Spent (Both Floors)") + 1
    time_col_letter = get_column_letter(time_col_idx)

    # Add "Remark" column header
    remark_col_idx = time_col_idx + 1
    remark_col_letter = get_column_letter(remark_col_idx)
    ws.cell(row=1, column=remark_col_idx).value = "Remark"

    # Add formula in each row
    for row in range(2, ws.max_row + 1):
        time_cell = f"{time_col_letter}{row}"
        remark_cell = f"{remark_col_letter}{row}"
        ws[remark_cell] = f'=IF({time_cell}="00:00:00", "00:00:00", TEXT(ABS(TIME(8,0,0) - TIMEVALUE({time_cell})), "hh:mm:ss"))'

    # Define fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Conditional formatting on the Remark column
    ws.conditional_formatting.add(
        f"{remark_col_letter}2:{remark_col_letter}{ws.max_row}",
        FormulaRule(formula=[f'AND(TIMEVALUE({time_col_letter}2)>=TIME(8,0,0), {time_col_letter}2<>"00:00:00")'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"{remark_col_letter}2:{remark_col_letter}{ws.max_row}",
        FormulaRule(formula=[f'AND(TIMEVALUE({time_col_letter}2)<TIME(8,0,0), {time_col_letter}2<>"00:00:00")'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"{remark_col_letter}2:{remark_col_letter}{ws.max_row}",
        FormulaRule(formula=[f'{time_col_letter}2="00:00:00"'], fill=gray_fill)
    )

    wb.save(excel_path)

    return render_template("merged.html")

##########################################################
app.secret_key = "your_secret_key"  # Replace with a secure key

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Admin credentials
ADMIN_USERNAME = "varun"
ADMIN_PASSWORD_HASH = generate_password_hash("varun123")  # Replace with a secure password

# Authentication decorator
def login_required(f):
    def wrapper(*args, **kwargs):
        if "admin_logged_in" not in session:
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

@app.route("/varun", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session["admin_logged_in"] = True
            return redirect(url_for("admin_portal"))
        else:
            return render_template("admin_login.html", error="SAD ADMIN TYPO WRONG 😈")

    return render_template("admin_login.html")

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login"))

@app.route("/admin/portal", methods=["GET"])
@login_required
def admin_portal():
    files = os.listdir(OUTPUT_FOLDER)
    exclude_filename = f"2nd_floor&3rd_floor_Combine_{current_date}.xlsx"

    files = [file for file in files if os.path.isfile(os.path.join(OUTPUT_FOLDER, file)) and file != exclude_filename]

    return render_template("admin_portal.html", files=files)


@app.route("/admin/download/<filename>")
@login_required
def admin_download(filename):
    return send_from_directory(OUTPUT_FOLDER, filename, as_attachment=True)

@app.route("/admin/delete/<filename>", methods=["POST"])
@login_required
def admin_delete(filename):
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(file_path):
        os.remove(file_path)
        return redirect(url_for("admin_portal"))
    return "File not found", 404

########################################## 3rd floor cal#####################################################################################

def calculate_total_time(punch):
    if pd.isna(punch) or punch.strip() == "":
        return "00:00"

    punch_times = re.findall(r'(\d{2}:\d{2})\((in|out)\)', punch)
    events = [(datetime.strptime(time, "%H:%M"), status) for time, status in punch_times]

    last_in = None
    total_time = 0
    i = 0

    while i < len(events):
        time, status = events[i]

        if status == "in":
            last_in = time
        elif status == "out" and last_in:
            while i + 1 < len(events) and events[i + 1][1] == "out":
                i += 1
            total_time += (events[i][0] - last_in).seconds
            last_in = None

        i += 1

    hours = total_time // 3600
    minutes = (total_time % 3600) // 60

    return f"{hours:02}:{minutes:02}"

@app.route("/3rdcal", methods=["GET", "POST"])
def index():
    total_time_spent = None

    if request.method == "POST":
        punch_record = request.form.get("punch_record", "")
        total_time_spent = calculate_total_time(punch_record)

    return render_template(".html", total_time_spent=total_time_spent)
##################################################################################################

@app.route('/', methods=['GET', 'POST'])
def hello():
    return render_template('hello.html')

@app.route('/welcome', methods=['GET', 'POST'])
def welcome():
    return render_template('welcome.html')

@app.route('/netflx', methods=['GET', 'POST'])
def netflx():
    return render_template('netflx.html')

if __name__ == "__main__":
    app.run(debug=True)
